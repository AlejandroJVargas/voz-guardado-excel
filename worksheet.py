import speech_recognition as sr
from openpyxl import load_workbook
from datetime import datetime


def iniciar_reconocimiento():
    reconocedor = sr.Recognizer()

    try:
        wb = load_workbook(
            "C:/Users/ale50/OneDrive/Escritorio/Pyhton/creador-excel-py/excel.xlsx"
        )
        ws = wb.active
    except FileNotFoundError:
        print("El archivo no se encontro. Se creara uno nuevo")
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Convertidor de voz"

        ws.append(["Fecha", "Hora", "Minuto", "Segundo", "Texto"])

    while True:
        try:
            with sr.Microphone() as mic:
                reconocedor.adjust_for_ambient_noise(mic)
                print(
                    "Ahora di unas palabras por un minuto, si quieres salir, solo di 'salir'"
                )

                audio = reconocedor.listen(mic, timeout=120, phrase_time_limit=120)
                texto = reconocedor.recognize_google(audio, language="es-ES")
                texto = texto.lower().strip()
                print("Estas son tus palabras -->", texto)

                ahora = datetime.now()
                fecha = ahora.strftime("%Y-%m-%d")
                hora = ahora.strftime("%H")
                minutos = ahora.strftime("%M")
                segundo = ahora.strftime("%S")

                ws.append([fecha, hora, minutos, segundo, texto])
                wb.save(
                    "C:/Users/ale50/OneDrive/Escritorio/Pyhton/creador-excel-py/excel.xlsx"
                )

                if "salir" == texto:
                    print("Haz terminado el programa")
                    break
        except sr.UnknownValueError:
            print(
                "Lo sentimos, no logramos comprender lo que dijiste. Inténtalo nuevamente."
            )
        except sr.RequestError:
            print("Lo sentimos, los servidores de Google no estan funcionando bien")
        except Exception as e:
            print(f"Lo sentimos pero ocurrio un error: {e}")


if __name__ == "__main__":
    iniciar_reconocimiento()
