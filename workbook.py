from openpyxl import load_workbook
# from openpyxl import Workbook

# wb = Workbook() # Sobre escribimos el archivo y no queremos eso

wb = load_workbook(
    "C:/Users/ale50/OneDrive/Escritorio/Pyhton/creador-excel-py/excel.xlsx"
)

print("Tu archivo fue cargado con exito")

ws = wb.active

# ws["A1"] = "Dia"
# ws["B1"] = "Hora"
# ws["C1"] = "Minuto"
# ws["D1"] = "Segundos"
# ws["E1"] = "Texto"

# ws.title = "Convertor de voz"


wb.save("C:/Users/ale50/OneDrive/Escritorio/Pyhton/creador-excel-py/excel.xlsx")

print("Se guardo con exito")
# titulo = wb.create_sheet("Texto")

# print(f"Titulo: {titulo}")
